import sys
import os
import time
import getopt
import threading
from PyQt5.QtWidgets import QApplication, QWidget, QHBoxLayout,QLineEdit, QProgressBar
from PyQt5.QtWidgets import QFileDialog, QTableWidget, QTableWidgetItem
from PyQt5.QtWidgets import QMessageBox, QPushButton
from PyQt5.QtWidgets import QGridLayout, QComboBox, QVBoxLayout, QLabel, QAbstractItemView
from PyQt5.QtCore import *
try:
    #pip install xlwt
    #pip install xlrd
    import xlwt, xlrd
except Exception as err:
    print("Import xlwt and xlrd {0}".format(err))

try:
    #pip install openpyxl
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, BarChart3D, PieChart, PieChart3D, LineChart
    from openpyxl.chart import ProjectedPieChart, Reference, Series
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
    from openpyxl.chart.series import DataPoint
    from openpyxl import load_workbook
    #from openpyxl.chart import View3D
except Exception as err:
    print("Import openpyxl {0}".format(err))

class XLSSplitter(QWidget):
    PUSHBUTTON_RUN = None
    PUSHBUTTON_OPEN_FILE = None
    LINEEDIT_FILENAME = None
    COMBOBOX_SHEETS = None
    TABLEWIDGET_MAIN_TABLE = None
    LABEL_INFORMATION = None
    PUSHBUTTON_SELECT_TARGET = None
    COMBOBOX_TARGET_COLUMNS = None
    LINEEDIT_TARGET_PATH = None
    PROGRESSBAR = None
    __XLS_FILE__ = None
    __SHEETNAME__ = None
    __SOURCE_COLUMNS__ = []
    __SOURCE_DATA__ = []
    __SELECT_GROUP__ = None
    __SELECT_PATH__ = None

    def __init__(self, file=None, encoding=None, path=None):
        super().__init__()
        if encoding is not None:
            self.__setting__['encoding'] = encoding
        if file is not None:
            self.__default_file__ = file
        if path is not None:
            self.__setting__['runpath'] = path

        self.initUI()

    def initUI(self):
        main_v_box = QVBoxLayout()

        file_grid = QGridLayout()
        file_grid.setSpacing(1)

        self.LINEEDIT_FILENAME = QLineEdit()
        self.LINEEDIT_FILENAME.setReadOnly(True)
        #LINEEDIT_FILENAME.setStyleSheet("min-height:22px;")
        self.PUSHBUTTON_OPEN_FILE = QPushButton("源文件")
        self.PUSHBUTTON_OPEN_FILE.clicked.connect(self.openFileXls)

        self.COMBOBOX_SHEETS = QComboBox()
        self.COMBOBOX_SHEETS.setStyleSheet("min-height:17px;")
        self.COMBOBOX_SHEETS.currentIndexChanged.connect(self.selectionchange)
        file_grid.addWidget(self.PUSHBUTTON_OPEN_FILE, 1, 0, 1, 1)
        file_grid.addWidget(self.LINEEDIT_FILENAME, 1, 1, 1, 8)
        file_grid.addWidget(self.COMBOBOX_SHEETS, 1, 9, 1, 3)

        self.TABLEWIDGET_MAIN_TABLE = QTableWidget()
        self.TABLEWIDGET_MAIN_TABLE.setObjectName("tableWidget")
        self.TABLEWIDGET_MAIN_TABLE.setColumnCount(0)
        self.TABLEWIDGET_MAIN_TABLE.setRowCount(0)
        self.TABLEWIDGET_MAIN_TABLE.setStyleSheet("selection-background-color:pink")
        self.TABLEWIDGET_MAIN_TABLE.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.TABLEWIDGET_MAIN_TABLE.setSelectionBehavior(QTableWidget.SelectRows)
        self.TABLEWIDGET_MAIN_TABLE.raise_()

        target_grid = QGridLayout()
        target_grid.setSpacing(1)
        self.COMBOBOX_TARGET_COLUMNS = QComboBox()
        self.COMBOBOX_TARGET_COLUMNS.currentIndexChanged.connect(self.groupChange)
        self.COMBOBOX_TARGET_COLUMNS.setStyleSheet("min-height:17px;")

        self.PUSHBUTTON_SELECT_TARGET = QPushButton("拆分到")
        self.PUSHBUTTON_SELECT_TARGET.clicked.connect(self.setTargetPath)
        #self.PUSHBUTTON_SELECT_TARGET.setStyleSheet("min-height:22px;")
        self.LINEEDIT_TARGET_PATH = QLineEdit()
        self.LINEEDIT_TARGET_PATH.setReadOnly(True)
        #dirname.setStyleSheet("min-height:22px;")
        target_grid.addWidget(self.COMBOBOX_TARGET_COLUMNS, 1, 0, 1, 2)
        target_grid.addWidget(self.PUSHBUTTON_SELECT_TARGET, 1, 2, 1, 1)
        target_grid.addWidget(self.LINEEDIT_TARGET_PATH, 1, 3, 1, 6)

        infor = QGridLayout()
        self.LABEL_INFORMATION = QLabel()
        self.LABEL_INFORMATION.setStyleSheet("color:brown;font-size:12px;font-weight:bold")
        infor.addWidget(self.LABEL_INFORMATION, 1, 0)

        hbox = QHBoxLayout()
        self.PROGRESSBAR = QProgressBar()
        hbox.addWidget(self.PROGRESSBAR)
        self.PUSHBUTTON_RUN = QPushButton("开始")
        self.PUSHBUTTON_RUN.clicked.connect(self.startSplit)
        helpButton = QPushButton("帮助")
        helpButton.clicked.connect(self.help)

        #hbox.addStretch(1)
        hbox.addWidget(self.PUSHBUTTON_RUN)
        hbox.addWidget(helpButton)

        main_v_box.addLayout(file_grid)
        main_v_box.addWidget(self.TABLEWIDGET_MAIN_TABLE)
        main_v_box.addLayout(target_grid)
        main_v_box.addLayout(infor)
        main_v_box.addLayout(hbox)

        self.setLayout(main_v_box)
        self.setGeometry(300, 300, 950, 550)
        self.setWindowTitle('XLS Splitter')
        self.show()

        self.help()

    def help(self):
        QMessageBox.information(self, '关于', '''<font size=18>XLS Splitter</font><br><hr><br>
                                        <font color=brown>欢迎使用.<br><br></font>
                                        <pre><p>{0}杨凯<br>{1}yangkai.bj@ccb.com</p></pre>'''.format('&#9;'*4, '&#9;'*4),
                                QMessageBox.Ok)

    def startSplit(self):
        split = threading.Thread(target=self.splitData(), args=(None,))
        split.start()

    def splitData(self):
        try:
            self.PUSHBUTTON_RUN.setEnabled(False)
            if self.__SELECT_GROUP__ is not None:
                index = self.__SELECT_GROUP__["index"]
                self.PROGRESSBAR.setMaximum(len(self.__SOURCE_DATA__))
                total = 0
                for group in self.__SELECT_GROUP__["content"].keys():
                    try:
                        subtotal = 0
                        target = os.path.abspath("{0}{1}{2}{3}".format(self.__SELECT_PATH__, os.path.sep, group, ".xlsx"))
                        self.LABEL_INFORMATION.setText("{0} ...".format(target))
                        workbook = Workbook()
                        sheet = workbook.active
                        sheet.title = group
                        row_index = 1
                        columns_font = Font(name='verdana',
                                            size=11,
                                            bold=True,
                                            italic=False,
                                            vertAlign=None,
                                            underline='none',
                                            strike=False,
                                            color='FFFAF0')
                        data_font = Font(name='verdana',
                                         size=11,
                                         bold=False,
                                         italic=False,
                                         vertAlign=None,
                                         underline='none',
                                         strike=False,
                                         color='FF000000')

                        colums_style = NamedStyle(name="colums_style")
                        colums_style.font = columns_font
                        colums_style.fill = PatternFill("solid", fgColor="8B0000")  # 背景填充
                        bd = Side(style='thick', color="A52A2A")
                        colums_style.border = Border(bottom=bd)
                        for col_index, col in enumerate(self.__SOURCE_COLUMNS__):
                            sheet.cell(row=row_index, column=col_index+1).value = str(col)
                            sheet.cell(row=row_index, column=col_index+1).style = colums_style
                            col_index += 1

                        data_style_a = NamedStyle(name="data_style_a")
                        data_style_a.font = data_font
                        data_style_a.fill = PatternFill("solid", fgColor="FAEBD7")  # 背景填充
                        bd = Side(style='thin', color="A52A2A")
                        data_style_a.border = Border(bottom=bd)
                        # left = bd, top = bd, right = bd, bottom = bd

                        data_style_b = NamedStyle(name="data_style_b")
                        data_style_b.font = data_font
                        data_style_a.fill = PatternFill("solid", fgColor="FFFAF0")  # 背景填充
                        bd = Side(style='thin', color="A52A2A")
                        data_style_b.border = Border(bottom=bd)
                        # left = bd, top = bd, right = bd, bottom = bd

                        row_index += 1

                        for count, row in enumerate(self.__SOURCE_DATA__):
                            if row[index] == group:
                                subtotal += 1
                                self.LABEL_INFORMATION.setText("{0} ... {1}".format(target, subtotal))
                                for col_index, col in enumerate(row):
                                    sheet.cell(row=row_index, column=col_index + 1).value = col
                                    if row_index % 2 == 0:
                                        sheet.cell(row=row_index, column=col_index + 1).style = data_style_b
                                    else:
                                        sheet.cell(row=row_index, column=col_index + 1).style = data_style_a
                                row_index += 1
                                total += 1
                                self.PROGRESSBAR.setValue(total)
                        workbook.save(filename=target)
                    except Exception as err:
                        QMessageBox.information(self, '错误', "CONTENT={0};ERROR={1}".format(group, err), QMessageBox.Ok)
            self.PUSHBUTTON_RUN.setEnabled(True)
        except Exception as err:
            QMessageBox.information(self, '错误', "ERROR={0}".format(err), QMessageBox.Ok)

    def setTargetPath(self):
        path = QFileDialog.getExistingDirectory(self, "选择", os.getcwd())
        if path != "":
            self.__SELECT_PATH__ = os.path.abspath(path)
            self.LINEEDIT_TARGET_PATH.setText(self.__SELECT_PATH__)

    def getGroup(self, index):
        group = {}
        for row in self.__SOURCE_DATA__:
            if row[index] not in group.keys():
                group[row[index]] = 1
            else:
                group[row[index]] += 1
        return group

    def groupChange(self):
        self.__SELECT_GROUP__ = {"index": self.COMBOBOX_TARGET_COLUMNS.currentIndex(), "content": self.getGroup(self.COMBOBOX_TARGET_COLUMNS.currentIndex())}
        self.LABEL_INFORMATION.setText("依据 {0} 将拆分为 {1} 个文件.".format(self.COMBOBOX_TARGET_COLUMNS.currentText(), len(self.__SELECT_GROUP__["content"].keys())))

    def selectionchange(self):
        self.__SHEETNAME__ = self.COMBOBOX_SHEETS.currentText()
        if os.path.splitext(self.__XLS_FILE__)[1].upper() == ".XLS":
            sheet, self.__SOURCE_COLUMNS__,self.__SOURCE_DATA__, error = self.readXLSData(self.__XLS_FILE__, self.__SHEETNAME__)
        elif os.path.splitext(self.__XLS_FILE__)[1].upper() == ".XLSX":
            sheet, self.__SOURCE_COLUMNS__, self.__SOURCE_DATA__, error = self.readXLSXData(self.__XLS_FILE__, self.__SHEETNAME__)
        self.showData(self.__SOURCE_COLUMNS__, self.__SOURCE_DATA__)

    def showData(self, columns, data):
        try:
            self.TABLEWIDGET_MAIN_TABLE.clear()
            self.TABLEWIDGET_MAIN_TABLE.setColumnCount(len(columns))
            self.TABLEWIDGET_MAIN_TABLE.setRowCount(len(data))
            self.TABLEWIDGET_MAIN_TABLE.setHorizontalHeaderLabels(columns)
        except Exception as err:
            QMessageBox.information(self, '错误', "ERROR={0}".format(err), QMessageBox.Ok)

        self.COMBOBOX_TARGET_COLUMNS.clear()
        for index, column in enumerate(columns):
            self.COMBOBOX_TARGET_COLUMNS.addItem(str(column), index)

        for i in range(len(data)):
            if i > 1000:
                QMessageBox.information(self, '注意', "数据超出1000条,超出部分不显示.", QMessageBox.Ok)
                break
            row = data[i]
            for j in range(len(row)):
                newItem = QTableWidgetItem(str(row[j]))
                newItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                self.TABLEWIDGET_MAIN_TABLE.setItem(i, j, newItem)
        self.LABEL_INFORMATION.setText("共计 {0} 条记录.".format(len(data)))

    def openFileXls(self):
        dig = QFileDialog(self, '打开...', None,
                          'Xls files (*.xls *.xlsx)')
        dig.setFileMode(QFileDialog.AnyFile)
        dig.setFilter(QDir.Files)
        if dig.exec_():
            self.__XLS_FILE__ = os.path.abspath(dig.selectedFiles()[0])
            self.LINEEDIT_FILENAME.setText(self.__XLS_FILE__)
            self.__SELECT_PATH__ = os.path.split(self.__XLS_FILE__)[0]
            self.LINEEDIT_TARGET_PATH.setText(self.__SELECT_PATH__)
            self.COMBOBOX_SHEETS.clear()
            self.LABEL_INFORMATION.setText("正在读取文件......")
            time.sleep(1)
            _sheets = self.getSheetNames(self.__XLS_FILE__)
            for sheet in _sheets:
                self.COMBOBOX_SHEETS.addItem(sheet)
            self.COMBOBOX_SHEETS.setCurrentText(_sheets[0])

    def getSheetNames(self, __XLS_FILE__):
        sheets = []
        if os.path.splitext(__XLS_FILE__)[1].upper() ==".XLS":
            wb = xlrd.open_workbook(__XLS_FILE__)
            for i, sheet in enumerate(wb.sheet_names()):
                sheets.append(sheet)

        elif os.path.splitext(__XLS_FILE__)[1].upper() == ".XLSX":
            wb = load_workbook(__XLS_FILE__, data_only=True)
            for i, sheet in enumerate(wb.sheetnames):
                sheets.append(sheet)
        return sheets

    def checkNONEData(self, value):
        if value is None:
            return ""
        else:
            return value

    def readXLSData(self, __XLS_FILE__, sheetname):
        #计数从1开始
        try:
            cols = []
            rows = []
            error = None
            wb = xlrd.open_workbook(__XLS_FILE__)
            ws = wb.sheet_by_name(sheetname)
            for i in range(ws.nrows):
                if i == 0:
                    cols = [self.checkNONEData(cell) for c, cell in enumerate(ws.row_values(i), start=0)]
                else:
                    rows.append([self.checkNONEData(cell) for c, cell in enumerate(ws.row_values(i), start=0)])
        except Exception as err:
            error = str(err)
        return sheetname, cols, rows, error

    def readXLSXData(self, xlsxfile, sheetname, row_start=1, row_end=1048576, col_start=1, col_end=16384):
        #计数从1开始
        try:
            cols = []
            rows = []
            error = None
            wb = load_workbook(xlsxfile, data_only=True)
            ws = wb[sheetname]
            for i, row in enumerate(ws.iter_rows(), start=0):
                if i == 0:
                    cols = [self.checkNONEData(cell.value) for c, cell in enumerate(row, start=0)]
                else:
                    rows.append([self.checkNONEData(cell.value) for c, cell in enumerate(row, start=0)])
        except Exception as err:
            error = str(err)
        return sheetname, cols, rows, error

if __name__ == '__main__':
    app = QApplication(sys.argv)
    file = None
    encoding = None
    path = None
    try:
        options, args = getopt.getopt(sys.argv, 'HhF:f:E:e:',
                                      ['HELP', 'help', 'FILE=', 'file=', 'ENCODING=', 'encoding='])

        for option, value in options:
            if str(option).upper() in ('-H', '--HELP'):
                print('Usage:')
                print('\t-f <FILE> -e <Encoding>')
                print('\t--file=<File> --Encoding=<Encoding>')
                print('\t<File> <Encoding>')
                sys.exit()
            if str(option).upper() in ('-F', '--FILE'):
                file = value
            if str(option).upper() in ('-E', '--ENCODING'):
                port = value
        for index, arg in enumerate(args):
            if file is None and index == 1:
                file = arg
            if encoding is None and index == 2:
                encoding = arg
            if index == 3:
                path = arg
    except Exception:
        pass

    ex = XLSSplitter(file=file, path=path)
    sys.exit(app.exec_())